from flask import Flask, render_template, request, redirect, url_for
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import base64
import os.path

app = Flask(__name__)

# Landing Page
@app.route('/')
def index():
    return render_template('index.html')

# Income Form
@app.route('/income', methods=['GET', 'POST'])
def income():
    if request.method == 'POST':
        name = request.form['name']
        date = request.form['date']
        service_type = request.form['service_type']
        amount = request.form['amount']
        transaction_type = request.form['transaction_type']
        reference_no = request.form['reference_no']

        # Append data to DataFrame
        data = {'Name': [name], 'Date': [date], 'Service Type': [service_type], 'Amount': [amount], 'Transaction Type': [transaction_type], 'Reference No.': [reference_no]}
        df = pd.DataFrame(data)
        
        # Write DataFrame to Excel file
        def file_contains_data(file_path):
          return os.path.exists(file_path) and os.path.getsize(file_path) > 0
        
        file_path = 'income.xlsx'

        if file_contains_data(file_path):
    
            try:
                existing_wb = load_workbook(file_path)
                existing_ws = existing_wb.active
                next_row = existing_ws.max_row + 1
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), next_row):
                    for c_idx, value in enumerate(row, 1):
                        existing_ws.cell(row=r_idx, column=c_idx, value=value)
                existing_wb.save(file_path)
                print("Data appended successfully to", file_path)
            except Exception as e:
                print("An error occurred:", e)
        else:
            # If the file doesn't exist or is empty, create a new file and save the data with headers
            df.to_excel(file_path, index=False)
            print("Created new file with data.")
        return redirect(url_for('index'))
    
    return render_template('income_form.html')

# Expenditure Form
@app.route('/expenditure', methods=['GET', 'POST'])
def expenditure():
    if request.method == 'POST':
        name = request.form['name']
        date = request.form['date']
        purpose = request.form['purpose']
        amount = request.form['amount']

        # Append data to DataFrame
        data = {'Name': [name], 'Date': [date], 'Purpose': [purpose], 'Amount': [amount]}
        df = pd.DataFrame(data)
        
        # Write DataFrame to Excel file
        def file_contains_data(file_path):
          return os.path.exists(file_path) and os.path.getsize(file_path) > 0
        file_path = 'expenditure.xlsx'

       
        if file_contains_data(file_path):
    
            try:
                existing_wb = load_workbook(file_path)
                existing_ws = existing_wb.active
                next_row = existing_ws.max_row + 1
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), next_row):
                    for c_idx, value in enumerate(row, 1):
                        existing_ws.cell(row=r_idx, column=c_idx, value=value)
                existing_wb.save(file_path)
                print("Data appended successfully to", file_path)
            except Exception as e:
                print("An error occurred:", e)
        else:
            # If the file doesn't exist or is empty, create a new file and save the data with headers
            df.to_excel(file_path, index=False)
            print("Created new file with data.")

        return redirect(url_for('index'))
    
    return render_template('expenditureform.html')

# Report Page
@app.route('/report')
def report():
    # Read Excel files
    income_df = pd.read_excel('income.xlsx')
    expenditure_df = pd.read_excel('expenditure.xlsx')

    # Calculate total income and expenditure
    total_income = income_df['Amount'].astype(float).sum()
    total_expenditure = expenditure_df['Amount'].astype(float).sum()
    profit_loss = total_income - total_expenditure

    # Bar chart for income and expenditure
    fig, ax = plt.subplots(figsize=(10, 6))

    income_df.groupby('Service Type')['Amount'].sum().plot(kind='bar', ax=ax, color='blue', alpha=0.7, label='Income')
    expenditure_df.groupby('Purpose')['Amount'].sum().plot(kind='bar', ax=ax, color='red', alpha=0.7, label='Expenditure')

    ax.set_title('Income vs Expenditure')
    ax.set_xlabel('Categories')
    ax.set_ylabel('Amount')
    ax.legend()

    # Save plot to a bytes object
    plt_bytes = io.BytesIO()
    plt.savefig(plt_bytes, format='png')
    plt_bytes.seek(0)
    plot_base64 = base64.b64encode(plt_bytes.getvalue()).decode()

    # Pie chart for profit and loss
    labels = ['Total Income', 'Total Expenditure']
    sizes = [total_income, total_expenditure]
    colors = ['#ff9999', '#66b3ff']
    fig1, ax1 = plt.subplots()
    ax1.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
    ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

    plt_bytes_pie = io.BytesIO()
    fig1.savefig(plt_bytes_pie, format='png')
    plt_bytes_pie.seek(0)
    pie_base64 = base64.b64encode(plt_bytes_pie.getvalue()).decode()

    # Render the report template with the tables and charts
    return render_template('reports.html', 
                           income_table=income_df.to_html(index=False), 
                           expenditure_table=expenditure_df.to_html(index=False),
                           total_income=total_income,
                           total_expenditure=total_expenditure,
                           profit_loss=profit_loss,
                           plot_base64=plot_base64,
                           pie_base64=pie_base64)

if __name__ == '__main__':
    app.run(debug=True)
